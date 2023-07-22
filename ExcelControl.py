from typing import Any

import openpyxl


class ExcelControl():

    def __init__(self, excel_name=None):

        self._workbook = openpyxl.Workbook()

        if excel_name:
            self._workbook = openpyxl.load_workbook(filename=excel_name)

        self._worksheet = self.workbook.active
        self._cell = self.worksheet.cell(row=1, column=1)

    @property
    def workbook(self) -> openpyxl.Workbook:
        return self._workbook

    @workbook.setter
    def workbook(self, workbook: openpyxl.Workbook) -> None:
        self._workbook = workbook

    @property
    def worksheet(self):
        return self._worksheet

    @worksheet.setter
    def worksheet(self, worksheet):
        self._worksheet = worksheet

    @property
    def cell(self):
        return self._cell

    @cell.setter
    def set_cell_value(self, value: Any | None):
        self.cell.value = value

    def print_cell_value(self) -> None:
        """
        print value of a cell
        """

        print(self.cell.value)

    def print_all(self) -> None:
        """
        print as
        
        [10, 'name', 'method'  ....]
        """
        for row in self.worksheet.iter_rows():
            print([row[i].value for i in range(len(row))])

    # def create_pretty_format(self):
    #     """
    #     create pretty format
    #     """

    #     wb_output =  openpyxl.Workbook()
    #     ws_output = wb_output.active



if __name__ == "__main__":
    ev = ExcelControl("data_sample.xlsx")
    # ev.set_cell_value(1)
    print(ev.workbook)
    ev.print_cell_value()
    ev.print_all()